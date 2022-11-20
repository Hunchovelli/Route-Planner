#Myron and Ademide

import openpyxl
import networkx as nx
from networkx import single_source_dijkstra, NodeNotFound, NetworkXNoPath
#imports module which gets the current day 
from datetime import datetime
#import module which draws line graph for the station densities
import matplotlib.pyplot as plt
import requests
import re
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timezone
from selenium.webdriver.chrome.options import Options

wrkbk = openpyxl.load_workbook("Stations_Updated.xlsx")  # Creating a link to the excel file with the London Underground Data
ws = wrkbk['Sheet1']

# This table will be used to create a dictionary with all the stations along with the lines that operate at that station
rows1 = ws.iter_rows(min_row=1, max_row=380, min_col=1, max_col=2, values_only=True)

# This table will be used to create the list of edges with weights needed for the network
rows2 = ws.iter_rows(min_row=1, max_row=377, min_col=5, max_col=7, values_only=True)

# Dictionary to store stations as keys and list of operating lines as values.
# Example: {'Green Park':['Picadilly', 'Jubilee', 'Victoria']}
station_dict = {}

# List of all the edges within our network
edges = []

# Adding the stations to dictionary along with all the lines that operate at the stations
for row in rows1:
    station = row[1].rstrip()  # Using rstrip() to get rid of unnecessary whitespace at end of string if present
    if station in station_dict.keys():
        station_dict[station].append(row[0])
        station_dict[station].sort()
    else:
        station_dict[station] = [row[0]]
        station_dict[station].sort()
    

# Get list of vertices by obtaining all the keys from the dictionary
vertices = station_dict.keys()

# print(vertices)  # Run this for a better idea

# Append all the edges from the excel file
for row in rows2:
    start = row[0].rstrip()
    end = row[1].strip()
    edges.append((start, end, row[2]))

# print(edges)   # Run this for a better idea

##########################################################################################################
# THESE 2 FOR LOOPS ARE NOT PART OF TASK 1
# THEY CAN BE USED AS REFERENCE FOR TASK 2
# Get all the edges linked to one station
# for i in edges:
#     if i[0] == "South Kensington" or i[1] == "South Kensington" or i[2] == "South Kensington":
#         print(i)

# # Remove an edge linked to the station
# for j in edges:
#     if j == ('Gloucester Road', 'South Kensington', 3):
#         edges.remove(j)
###########################################################################################################

# CONTINUATION OF TASK 1

# Create the graph by passing the list of nodes and edges
G = nx.MultiGraph()
G.add_nodes_from(vertices)
G.add_weighted_edges_from(edges)


################ Processing Data for Task 3 ###############################

# Processing data from the Weekly Crowd Traffic spreadsheet
wrkbk2 = openpyxl.load_workbook("WeeklyCrowdTraffic_Updated.xlsx")  # Link to underground crowd traffic data
ws2 = wrkbk2['Sheet1']

rows3 = ws2.iter_rows(min_row=3, max_row=269, min_col=1, max_col=7, values_only=True)

# This dictionary will hold each station as keys and each key will hold its own dictionary of crowd density
# Example: station_density["Green Park"] ------>  {'Weekday': 63355, 'Saturday': 38963, 'Sunday': 27357}
# The numbers in the dictionary are the average number of people at the station on that given day
station_density = {}

for row in rows3:
    station = row[0].rstrip()
    if station in vertices:

        weekday_avg = (row[1] + row[4])/2  # Calculate the average number of people visiting the station on weekdays
        saturday_avg = (row[2] + row[5])/2  # Calculate the average number of people visiting the station on saturdays
        sunday_avg = (row[3] + row[6])/2  # Calculate the average number of people visiting the station on sundays

        # Add the averages to the dictionary
        station_density[station] = {}
        station_density[station]["Weekday"] = int(weekday_avg)
        station_density[station]["Saturday"] = int(saturday_avg)
        station_density[station]["Sunday"] = int(sunday_avg)
#################################################################################################################

# FUNCTION TO GENERATE URLs TO THE TFL WEBPAGES FOR EACH STATION

# Takes a list
def generate_url(list_of_stations):

    url = "https://tfl.gov.uk/disambiguation?Input=Hounslow+West&DataSetsJson=%5B%5B%22stopPoints%22%2C%22%2F%7B%7Bmode%7D%7D%2Fstop%2F%7B%7BnaptanId%7D%7D%2F%7B%7BstopName%7D%7D%2F%22%5D%2C%5B%22routes%22%2C%22%2F%7B%7Bmode%7D%7D%2Froute%2F%7B%7BlineIds%7D%7D%2F%22%5D%5D&Modes=tube&PlaceHolderText=Tube+station+or+line+%28e.g.+Victoria%29"

    links = []

    for i in list_of_stations:
        station_string = i + " Underground Station"
        response = requests.post(url, data={"input": station_string})
        links.append(response.url)
    
    # List of URLs for each station
    return links 


###############################################################################################################

# Task 1a
# Main program which will be take the input from user and display route
def get_route():

    # Count the number of times the user will need to change lines
    # Add 5 min to the total time for each line change
    line_change_counter = 0

    print("Enter the departure stations: ")
    departure = input()
    print("Enter the destination stations: ")
    destination = input()
    print("\n Your generated route is:\n")
    
    # Handle exceptions occurred from bad input
    try:
        route = single_source_dijkstra(G, source=departure, target=destination, weight='weight')
        print (route)
        # Print route, station by station along with the line they will use to get there
        for i in route[1]:

            # First station will not have an associated line because we assume that user is at the station
            if route[1].index(i) == 0:
                prev_station = i  # Holds the station travelled from
                prev_line = station_dict[i][0]  # Holds the line used to travel to the previous station
                print("Take the {} line from {}\n".format(prev_line, prev_station))
                continue

            chosen_line = ""

            # Determine which line the user will take to the new station
            for line in station_dict[i]:
                if line in station_dict[prev_station]:
                    chosen_line = line

            if chosen_line != prev_line:
                print("\nSwitch To ----> {} line\n".format(chosen_line))
                print("{} : via the {} line".format(i, chosen_line))
            else:
                print("{} : via the {} line".format(i, chosen_line))
                
                
            # Add to the line change counter if line change detected
            if chosen_line != prev_line:
                line_change_counter += 1

            # Update the previous station and line used
            prev_station = i
            prev_line = chosen_line

        # Add 30 seconds passenger boarding/disembarking time for each station excluding start and end stations
        station_halt_time = ((len(route[1]) - 2) * 30) / 60

        # Add a 5 minute delay each time the user has to switch to another line
        line_change_delay = line_change_counter * 5

        # Calculate the total time for the whole journey including delays
        total_time = int(route[0] + station_halt_time + line_change_delay)

        print("\nTotal route time including {} line changes is {} minutes".format(line_change_counter, total_time))

        #################################################################################################################

        #DRAWS LIVE GRAPH OF STATION STATUS

        utc_dt = datetime.now(timezone.utc)
        dt = utc_dt.astimezone().time()

        #list which will hold the x values for each station for the graph
        xaxis = []

        #for loop which iterates throught the list that shows the route
        for station in route[1]:
            #for each station in the route list we add it to the x axis list
            xaxis.append(station)

        #list which will hold the y values for each station on the station density for the graph
        yaxis = []

        #loops through the list of links for the status of the station for the route
        for link in generate_url(route[1]):
            
            #variable that holds the link for the status of station from the tfl website
            url = link
            #object that is used to set the capability and customization and configuration of the ChromeDriver session. 
            option = webdriver.ChromeOptions()
            #headless browser can access any website but unlike normal browsers nothing will appear on the screen. works in the backend and is invisble to user.
            option.add_argument('headless')
            #disable chrome logging output in the terminal
            option.add_experimental_option('excludeSwitches', ['enable-logging'])
            #object for the chrome browser to open
            driver = webdriver.Chrome(options=option)
            #opens the browser with the url page open
            driver.get(url)

            #initialize global dictionary that can be accessed outside the function
            global timevalue_dict
            #dictionary that stores the time with the corresponding value of the status
            timevalue_dict = {}

            #function that creates a dictionary with all the stored times and their corresponding values for the status of the station   
            def status_dict(time_range, value):
                #for loop that loops through the 
                for time in time_range:
                    #
                    timevalue_dict[time] = value
                #returns the dictionary
                return timevalue_dict

            #try block that
            try:
                #looks for the all the 'g' element in the html and stores them all in the tag. Waits a 100sec max for finding the element.
                tags = WebDriverWait(driver, 100).until(EC.presence_of_all_elements_located((By.TAG_NAME, 'g')))
                #creates empty list
                values_List = []

                #loops through each tag made
                for tag in tags:
                    #gets the value in aria-describedby and stores ti in value variable
                    value = tag.get_attribute('aria-describedby')
                    #checks to see if attribute is present in the g element.
                    if value != None:
                        #adds the value from the attribute to the list
                        values_List.append(value)
                
                #loops through the values in the value_list
                for value in values_List: 
                    #splits each value by in the string by numerical digits and creates a list of the split characters in the string
                    split_string = re.split('(\d+)', value)
                    #pass the split string elements through indexing the split_string list into the function declared earlier that creates a dictionary of the times and the corresponding values.
                    status_dict(range(int(split_string[1]+split_string[3]),(int(split_string[5]+split_string[7]))), int(split_string[9]))
            #after the try block has been executed
            finally:
                #quit the web scraping 
                driver.quit()

            #split the current time by : and puts it into a list
            live_time_split = str(dt).split(':')
            #merging the hour and minute of the time into one number
            merged_time = int(live_time_split[0]+live_time_split[1])
            #use the merged time to use as a key for the timevalue dictionary which will add the corresponding value to the y axis list.
            yaxis.append(timevalue_dict[merged_time])

        #plots the axis of the line graph
        plt.plot(xaxis, yaxis)
        #titles the line graph
        plt.title("Live Crowd/Traffic for Route at each Stations")
        #labels the x axis
        plt.xlabel("Station")
        #labels the y axis
        plt.ylabel("Current Crowd Traffic")
        #displays the produced graph
        plt.show()
    #################################################################################################################
    except(NodeNotFound, NetworkXNoPath):
        print("Invalid Nodes Entered")
    
    
get_route()





